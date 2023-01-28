import pandas as pd
import numpy as np
from datetime import datetime
from datetime import date
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment

def calculate_work_hours(row):
    time_delta = row["scan_pulang"] - row["scan_masuk"]
    time_delta = str(time_delta)
    hours = int(time_delta[7:9])
    minutes = int(time_delta[10:12])
    if minutes >= 50:
        hours += 1
    elif minutes >= 20:
        hours += 0.5

    if hours <= 8:
        jam_kerja = hours
        jam_lembur = 0
    elif hours > 8:
        jam_kerja = 8
        jam_lembur = hours - 8
    return jam_kerja, jam_lembur, time_delta

def calculate_salary(row):
    if  row["Tanggal"].weekday() == 6 or row["is_holiday"] == "Y": # tambahin kondisi kalo hari libur
        gaji_harian = (row["jam_kerja"]/8) * (row["Gaji Harian (Pokok)"]*1.5)
        gaji_lembur = row["jam_lembur"]* (row["Upah Lembur"]*1.5)
    else:    
        gaji_harian = (row["jam_kerja"]/8) * row["Gaji Harian (Pokok)"]
        gaji_lembur = row["jam_lembur"]*row["Upah Lembur"]
    total_gaji_harian = (gaji_harian + gaji_lembur + row["uang_makan_harian"]) - (row["denda_tidak_scan_masuk"]+row["denda_tidak_scan_pulang"])
    return gaji_harian, gaji_lembur, total_gaji_harian


def generate_kwitansi(row):
    file_list = []
    for idx in row.index:
        wb = load_workbook("Template Kwitansi.xlsx")
        sheet = wb.active

        #reading specific column
        # B3 - Nama
        sheet.cell(row=3, column=2).value=row["Nama"][idx]
        sheet.merge_cells('B3:I6')
        sheet.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        # H10 & E24 - gaji_final
        sheet.cell(row=10, column=8).value=row["gaji_final"][idx]
        sheet.cell(row=10, column=8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('H10:J10')  
        sheet.cell(row=24, column=5).value=row["gaji_final"][idx]
        sheet.merge_cells('E24:K24')
        sheet.cell(row=24, column=5).alignment = Alignment(horizontal='left', vertical='center')
        # G28 - Nama Bank & Nama Akun Bank
        sheet.cell(row=28, column=7).value=row["Nama Bank"][1] + " A/n "+ row["Nama Akun Bank"][1]
        # G30 - Nomor Rekening
        sheet.cell(row=30, column=7).value=row["Nomor Rekening"][idx]
        # V5 & G32 - tanggal dicetak
        sheet.cell(row=5, column=22).value= date.today().strftime('%d %b %Y')
        sheet.cell(row=32, column=7).value= date.today().strftime('%d %b %Y')
        # V3 - Nomor Kwitansi
        sheet.cell(row=3, column=22).value= "KWT_ATS_VII-22_"
        # M14 & R14 - Periode Upah
        sheet.cell(row=14, column=13).value=row["start_date"][idx]
        sheet.cell(row=14, column=18).value=row["end_date"][idx]
        wb.title = row["nama_worksheet"][idx]
        sheet.title = row["nama_worksheet"][idx]
        file_name = "Kwitansi_"+row["nama_worksheet"][idx]+"_"+str(row["start_date"][idx].strftime('%d%b'))+"-"+str(row["end_date"][idx].strftime('%d%b%Y'))+".xlsx"
        wb.save("kwitansi_output/"+file_name)
        file_list.append("kwitansi_output/"+file_name)
    return file_list